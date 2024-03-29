'''
tkauto.py
Python console program
Author: Michael Leidel
Description:
Builds Python tkinter starter code from an xlsx file.

usage: tkauto.py [-h] [-x] [-t] [-b] filename

positional arguments:
    filename    Excel or "nofile" for -t option

options:
  -h, --help  show this help message and exit
  -x          Execute with python3 after compile
  -t          Output just the template - layout="nofile"
  -b          Use ttkbootstrap template: tkbauto_tpl.py
'''
import os
import sys
import argparse
import openpyxl

outFile = "output.py"  # default output file name

# tkauto_tpl.py & tkbauto_tpl.py code templates
#   must be in same directory as this script
# This sets the path.
p = os.path.realpath(__file__)
TPLPATH = os.path.dirname(p) + "/"

parser = argparse.ArgumentParser(description='tkauto build Python tkinter GUI')
parser.add_argument('-x', dest='exec', action='store_true',
                    help='Execute with python3 after compile')
parser.add_argument('-t', dest='template', action='store_true',
                    help='Output just the template - layout="nofile"')
parser.add_argument('-b', dest='bstrap', action='store_true',
                    help='Use ttkbootstrap template: tkbauto_tpl.py')
parser.add_argument('filename', help='.xls_ | .odf, or "nofile" for -t option')
args = parser.parse_args()

# list (flds) index names for the (row)column values
nop, wgt, par, var, txt, com, row, col, rsp, csp, sty, owa = 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11

# the following used for error messages
colname = ['nop', 'wgt', 'parent', 'object variable', 'text', 'command', 'row', 'col']

sout = ""  # holds output for injection into template

errors = False  # detected missing required widget option(s)

# Menu Only: list (flds) index names for the (row)column values
nop, mid, lbl, cmd, acc, und = 0, 1, 2, 3, 4, 5
MN_VAR = ""  # holds the menu object variable found in column 1 os ss

def prt(s):
    ''' Adjust leading tabs/spaces for output here '''
    global sout
    sout += "        " + s + "\n"


def getRowSpan(val):
    ''' format rowspan option '''
    rowspan = ""
    if val != "":
        rowspan = ", rowspan=" + str(val)
    return rowspan


def getColSpan(val):
    ''' format colspan option '''
    colspan = ""
    if val != "":
        colspan = ", columnspan=" + str(val)
    return colspan


def getSticky(val):
    ''' format sticky option '''
    sticky = ""
    if val != "":
        sticky = ", sticky=" + "'" + val + "'"
    return sticky

def getAttrib(val):
    ''' format misc options '''
    atts = ""
    if val != "":
        # spreadsheets don't always use plain text
        atts = ", " + val.replace("’", "'").replace('”', '"')
    return atts

def checkcols(fields):
    ''' check input for required values in the spreadsheet columns
        each widget will have different required options '''
    global errors
    for item in fields:
        if flds[item] == "":
            print(f">>>>>>>>> {flds[1]} widget missing required option: {colname[item]}")
            errors = True
            return


# PROCESS MENU ITEMS

def proc_menu_item():
    ''' Process one line of menu code '''
    global MN_VAR
    global domenu

    print("proc_menu_item row ", flds[mid])

    if flds[mid].lower() == "endmenu":
        prt("root.config(menu=menubar) # display the menu\n\n")  # here ends the menu code
        domenu = False
        return

    ROUT = ""  # holds output for each row of python/tkinter code

    if flds[mid] != MN_VAR:  # starts a new menu or submenu item
        if flds[mid] == "submenu":
            ROUT = "submenu = Menu(" + MN_VAR +", tearoff=False)"
            MN_VAR = flds[mid]
            prt(ROUT)
            # now continue to process the row
        elif MN_VAR == "submenu":  # finished with this submenu?
            ROUT = flds[mid] + ".add_cascade(label=\""
            ROUT += flds[lbl][1:] + "\", menu=submenu"
            if flds[und] != "":
                ROUT += ", underline=" + flds[und]
            ROUT += ")"
            prt(ROUT)
            MN_VAR = flds[mid]
            return
        else:
            ROUT = flds[mid] + " = Menu(menubar, tearoff=0)"
            MN_VAR = flds[mid]
            prt(ROUT)
            # now return to process the row

    if flds[lbl] == "separator":
        ROUT = flds[mid] + ".add_separator()"
        prt(ROUT)
        return

    if flds[lbl].startswith("_"):  # end of children this is the top label
        ROUT = "menubar.add_cascade(label=\""
        ROUT += flds[lbl][1:] + "\", menu=" + flds[mid] + ")"
        prt(ROUT)
        return

    # VAR + '.add_command(label="LBL", command=self.CMD, accelerator="ACC", underline=UND)'
    ROUT = flds[mid] + ".add_command(label=\""
    ROUT += flds[lbl] + "\", command=self."
    ROUT += flds[cmd]
    callbacks.append(flds[cmd])
    if flds[acc] != "":
        ROUT += ", accelerator=\"" + flds[acc] + "\""
    if flds[und] != "":
        ROUT += ", underline=" + flds[und]
    ROUT += ")"
    prt(ROUT)

    # END proc_menu_item

'''
    Did the user request just the template?
'''
if args.template:
    fout = open(outFile, "w")

    # location of master templates
    if args.bstrap:
        fin = open(TPLPATH + "tkbauto_tpl.py", "r")
    else:
        fin = open(TPLPATH + "tkauto_tpl.py", "r")

    for line in fin:
        if line.find("INSERT TKAUTO OUTPUT") > 0:
            continue
        else:
            fout.write(line)
    fout.close()
    fin.close()
    sys.exit()


# verify the input xlsx file
filetype = ""  # .xls? or .ods

if not os.path.exists(args.filename) or not args.filename.endswith(".xlsx"):
    print("xlsx file missing or incorrect type: " + args.filename)
    sys.exit()
'''
    Get the workbook and spreadsheet.
'''
wb = openpyxl.load_workbook(args.filename)
# sheet = wb.get_sheet_by_name('layout')  # Must be a Sheet titled 'layout' !
sheet = wb.worksheets[0]  # first worksheet in the workbook
rownum = 3

flds = []
callbacks = []
domenu = False

#

print("TkAuto Starting -------------------------------------->")

while(True):

    # EOF: nothing in col 1 end the loop (very big loop)
    if sheet.cell(row=rownum, column=1).value is None:
        break

    if sheet.cell(row=rownum, column=1).value == "#":
        rownum += 1  # comments: increment row for loop
        continue

    flds.clear()  # clear list
    flds.append("nop")  # zero element not used

    # load up the flds list with this row's columns values
    for c in range(1, 12):  # columns align with list index
        val = sheet.cell(row=rownum, column=c).value
        if val is None:
            flds.append("")
        else:
            flds.append(val)

    # first check if processing menu
    if domenu is True:
        proc_menu_item()
        rownum += 1
        continue

    if flds[wgt].lower() == "begmenu":  # going to process a menu
        domenu = True
        prt("menubar = Menu(root)")  # the menu code starts here
        rownum += 1
        continue  # go back to process for "domenu"

    # Process this row/columns values

    print(flds[wgt])

    # These will be set to the formatted code or ""
    attribs = getAttrib(flds[owa])  # Other Widget Attributes
    rowspan = getRowSpan(flds[rsp])
    colspan = getColSpan(flds[csp])
    sticky = getSticky(flds[sty])


    # BUTTON
    if flds[wgt].lower() == "button":
        checkcols([2,3,4,5,6,7])
        line = "{0} = Button({4}, text='{1}', command=self.{2}{3})"
        prt(line.format(flds[var], flds[txt], flds[com], attribs, flds[par]))
        callbacks.append(flds[com])
        line = "{0}.grid(row={1}, column={2}{3}{4}{5}, padx=4, pady=4)\n"
        prt(line.format(flds[var], flds[row],
                        flds[col], rowspan, colspan, sticky))

    # LABEL
    elif flds[wgt].lower() == "label":
        checkcols([2,3,6,7])
        if flds[com] != "":  # a StringVar was given
            prt("self." + flds[com] + " = StringVar()")

            line = "{0} = Label({3}, textvariable=self.{1}{2})"
            prt(line.format(flds[var], flds[com], attribs, flds[par]))

            line = "{0}.grid(row={1}, column={2}{3}{4}{5}, padx=4, pady=4)"
            prt(line.format(flds[var], flds[row],
                            flds[col], rowspan, colspan, sticky))

            line = "self.{}.set('{}')\n"
            prt(line.format(flds[com], flds[txt]))
        else:  # without StringVar
            line = "{0} = Label({3}, text='{1}'{2})"
            prt(line.format(flds[var], flds[txt], attribs, flds[par]))
            line = "{0}.grid(row={1}, column={2}{3}{4}{5}, padx=4, pady=4)\n"
            prt(line.format(flds[var], flds[row],
                            flds[col], rowspan, colspan, sticky))

    # ENTRY
    elif flds[wgt].lower() == "entry":
        checkcols([2,3,5,6,7])
        prt("self." + flds[com] + " = StringVar()")
        prt("# self." + flds[com] + ".trace(\"w\", self.eventHandler)")
        line = "{0} = Entry({3}, textvariable=self.{1}{2})"
        prt(line.format(flds[var], flds[com], attribs, flds[par]))
        line = "{0}.grid(row={1}, column={2}{3}{4}{5}, padx=4, pady=4)\n"
        prt(line.format(flds[var], flds[row],
                        flds[col], rowspan, colspan, sticky))

    # TEXT
    elif flds[wgt].lower() == "text":
        checkcols([2,3,6,7])
        line = "self.{0} = Text({2}{1})"
        prt(line.format(flds[var], attribs, flds[par]))
        line = "self.{0}.grid(row={1}, column={2}{3}{4}{5})\n"
        prt(line.format(flds[var], flds[row],
                        flds[col], rowspan, colspan, sticky))
        line = '''
        # efont = Font(family="Helvetica", size=14)
        # self.EDITOR.configure(font=efont)
        # self.EDITOR.config(wrap="word", # wrap=NONE
        #                    undo=True, # Tk 8.4
        #                    width=50,
        #                    height=12,
        #                    padx=5, # inner margin
        #                    insertbackground='#000',   # cursor color
        #                    tabs=(efont.measure(' ' * 4),))
        # self.EDITOR.focus()
        ## basic Text widget commands #
        # .get("1.0", END)
        # .delete("1.0", END)
        # .insert("1.0", "New text content ...")
        ## Select All
        # .tag_add(SEL, '1.0', END)
        # .mark_set(INSERT, '1.0')
        # .see(INSERT)

        '''
        prt(line + "\n")

    # LIST
    elif flds[wgt].lower().startswith("list"):
        checkcols([2,3,6,7])
        line = "self.{0} = Listbox({2}{1}, exportselection=False)"
        prt(line.format(flds[var], attribs, flds[par]))
        rowspan = getRowSpan(flds[rsp])
        line = "self.{0}.grid(row={1}, column={2}{3}{4}{5})"
        prt(line.format(flds[var], flds[row],
                        flds[col], rowspan, colspan, sticky))

        line = '''
        # self.LISTBOX.bind("<<ListboxSelect>>", self.on_select_list)
        # for i in range(100):
        #     self.LISTBOX.insert(END, "Item " + str(i))'''
        prt(line)

        line = '''
    # HANDLER FOR LIST SELECTION
    # def on_select_list(self, event):
    #     list_item = self.LISTBOX.get(ANCHOR)
    #     list_inx = self.LISTBOX.index(ANCHOR)
    #     print(list_item, str(list_inx) +
    #           " of " + str(self.LISTBOX.size()))
    #
    # HANDLERS FOR EDITING LISTBOX
    #
    # def delete_item(self):
    #     if self.LISTBOX.curselection() == ():
    #         return  # nothing selected
    #     print("Deleting: " + self.LISTBOX.get(ANCHOR))
    #     self.LISTBOX.delete(self.LISTBOX.index(ANCHOR))
    #
    # def insert_item(self):
    #     if self.LISTBOX.curselection() == ():
    #         return  # nothing selected
    #     list_item = self.LISTBOX.get(ANCHOR)
    #     list_inx = self.LISTBOX.index(ANCHOR)
    #     self.LISTBOX.insert(list_inx, "Inserted this")
    #     print("inserted at " + str(list_inx)) '''
        prt(line + "\n")


    # VERT SCROLLBAR
    elif flds[wgt].lower() == "scrolly":
        checkcols([2,3,5,6,7])
        line = "self.{0} = Scrollbar({2}, orient=VERTICAL, command=self.{1}.yview)"
        if flds[com] == "":
            print("\n\nMISSING list object FOR SCROLLBAR WIDGET\n\n")
            sys.exit()
        prt(line.format(flds[var], flds[com], flds[par]))
        line = "self.{0}.grid(row={1}, column={2}{3}{4}{5})  # use nse"
        prt(line.format(flds[var], flds[row],
                        flds[col], rowspan, colspan, sticky))
        line = "self.{0}['yscrollcommand'] = self.{1}.set\n"
        prt(line.format(flds[com], flds[var]))

    # HORZ SCROLLBAR
    elif flds[wgt].lower() == "scrollx":
        checkcols([2,3,5,6,7])
        line = "self.{0} = Scrollbar({2}, orient=HORIZONTAL, command=self.{1}.xview)"
        if flds[com] == "":
            print("\n\nMISSING list/text object FOR SCROLLBAR WIDGET\n\n")
            sys.exit()
        prt(line.format(flds[var], flds[com], flds[par]))
        line = "self.{0}.grid(row={1}, column={2}{3}{4}{5})"
        prt(line.format(flds[var], flds[row],
                        flds[col], rowspan, colspan, sticky))
        line = "self.{0}['xscrollcommand'] = self.{1}.set\n"
        prt(line.format(flds[com], flds[var]))

    # CHECK BOX Checkbutton
    elif flds[wgt].lower().startswith("check"):
        checkcols([2,3,4,5,6,7])
        prt("self.{0} = IntVar()".format(flds[com]))
        line = "{0} = Checkbutton({4}, variable=self.{1}, text='{2}'{3})"
        prt(line.format(flds[var], flds[com], flds[txt], attribs, flds[par]))
        line = "{0}.grid(row={1}, column={2}{3}{4}{5}, padx=4, pady=4)\n"
        prt(line.format(flds[var], flds[row],
                        flds[col], rowspan, colspan, sticky))
    # RADIO BUTTON
    elif flds[wgt].lower().startswith("radio"):
        checkcols([2,3,4,5,6,7])
        prt("self." + flds[com] +
            " = StringVar() # USE ONE VAR PER GROUP OF BUTTONS")
        line = "{0} = Radiobutton({5}, variable=self.{1}, value='{2}', text='{2}'{4})"
        prt(line.format(flds[var], flds[com], flds[txt], flds[txt], attribs, flds[par]))
        line = "{0}.grid(row={1}, column={2}{3}{4}{5}, padx=4, pady=4)\n"
        prt(line.format(flds[var], flds[row],
                        flds[col], rowspan, colspan, sticky))
    # SPIN BOX
    elif flds[wgt].lower().startswith("spin"):
        checkcols([2,3,5,6,7])
        prt("self." + flds[com] + " = StringVar(value=0)")
        line = "{0} = Spinbox({3}, textvariable=self.{1}, from_=0, to=10{2})"
        prt(line.format(flds[var], flds[com], attribs, flds[par]))
        line = "{0}.grid(row={1}, column={2}{3}{4}{5}, padx=4, pady=4)\n"
        prt(line.format(flds[var], flds[row],
                        flds[col], rowspan, colspan, sticky))
    # OPTION MENU
    elif flds[wgt].lower().startswith("option"):
        checkcols([2,3,5,6,7])
        prt("optionlist = ('aaa', 'bbb', 'ccc', 'ddd', 'eee', 'fff')")
        prt("self." + flds[com] + " = StringVar()")
        prt("self." + flds[com] + ".set(optionlist[0])")
        line = "{0} = OptionMenu({2}, self.{1}, *optionlist)"
        prt(line.format(flds[var], flds[com], flds[par]))
        line = "{0}.grid(row={1}, column={2}{3}{4}{5}, padx=4, pady=4)\n"
        prt(line.format(flds[var], flds[row],
                        flds[col], rowspan, colspan, sticky))

    # COMBOBOX (from tkinter.ttk import Combobox)
    elif flds[wgt].lower().startswith("combo"):
        checkcols([2,3,5,6,7])
        prt("self." + flds[com] + " = StringVar()")
        line = "{0} = Combobox({3}, textvariable=self.{1}{2})"
        prt(line.format(flds[var], flds[com], attribs, flds[par]))
        line = "{0}['values'] = ('value1', 'value2', 'value3')"
        prt(line.format(flds[var]))
        prt("# COMBO.bind('<<ComboboxSelected>>', self.ONCOMBOSELECT)")
        prt("%s.current(0)" % flds[var])
        line = "{0}.grid(row={1}, column={2}{3}{4}{5}, padx=4, pady=4)\n"
        prt(line.format(flds[var], flds[row],
                        flds[col], rowspan, colspan, sticky))

    # PROGRESSBAR
    elif flds[wgt].lower().startswith("prog"):
        checkcols([2,3,6,7])
        line = "{0} = Progressbar({2}, orient='horizontal', mode='indeterminate', maximum=20 {1})"
        prt(line.format(flds[var], attribs, flds[par]))
        line = "{0}.grid(row={1}, column={2}{3}{4}{5})"
        prt(line.format(flds[var], flds[row],
                        flds[col], rowspan, colspan, sticky))
        line = ("# {0}.start() | {0}.stop() | {0}.grid_forget() | {0}.grid(self, ...) \n")
        prt(line.format(flds[var], flds[var], flds[var], flds[var]))

    # NOTEBOOK
    elif flds[wgt].lower().startswith("notebook"):
        checkcols([2,3,4,5,6,7])
        line = "{0} = Notebook({2}{1})"
        prt(line.format(flds[var], attribs, flds[par]))
        line = "tab1 = Frame({0}, width=99, height=99)  # need W & H"
        prt(line.format(flds[var]))
        line = "tab2 = Frame({0}, width=99, height=99)"
        prt(line.format(flds[var]))
        line = "tab3 = Frame({0}, width=99, height=99)"
        prt(line.format(flds[var]))
        line = "{0}.add(tab1, text='tab 1')"
        prt(line.format(flds[var]))
        line = "{0}.add(tab2, text='tab 2')"
        prt(line.format(flds[var]))
        line = "{0}.add(tab3, text='tab 3')"
        prt(line.format(flds[var]))
        line = "{0}.grid(row={1}, column={2}{3}{4}{5})\n"
        prt(line.format(flds[var], flds[row],
                        flds[col], rowspan, colspan, sticky))

    # FRAMES
    elif flds[wgt].lower() == "frame":
        checkcols([2,3,6,7])
        line = "{0} = Frame({2}{1})"
        prt(line.format(flds[var], attribs, flds[par]))
        line = "{0}.grid(row={1}, column={2}{3}{4}{5})"
        prt(line.format(flds[var], flds[row],
                        flds[col], rowspan, colspan, sticky))
        line = '''
        # lframe = LabelFrame(self, text="text",
        #                     width=100, height=100)
        # lframe.grid(row=1, column=1, sticky='nsew')
        #
        '''
        prt(line + "\n")


    # SEPARATOR
    elif flds[wgt].lower() == "separator":
        checkcols([2,3,6,7])
        line = "{0} = Separator({2}{1})"
        prt(line.format(flds[var], attribs, flds[par]))
        line = "{0}.grid(row={1}, column={2}{3}{4}{5})\n"
        prt(line.format(flds[var], flds[row],
                        flds[col], rowspan, colspan, sticky))

    # MESSAGEBOX
    elif flds[wgt].lower() == "messagebox":
        line = '''
        # from tkinter import messagebox
        # messagebox.showerror("Error", "Error message")
        # messagebox.showwarning("Warning", "Warning message")
        # messagebox.showinfo("Information", "Informative message")
        # messagebox.askokcancel("Message title", "Message content")
        # messagebox.askretrycancel("Message title", "Message content")
        #     ok, yes, retry returns TRUE
        #     no, cancel returns FALSE
        '''
        prt(line + "\n")

    # DIALOG (simpledialog)
    elif flds[wgt].lower() == "simpledialog":
        line = '''
        # # from tkinter import simpledialog
        # simpledialog.askfloat(title, prompt)
        # simpledialog.askinteger(title, prompt)
        # simpledialog.askstring(title, prompt)
        # if answer is not None:
        '''
        prt(line + "\n")

    # FILEDIALOG
    elif flds[wgt].lower().startswith("file"):
        line = '''
        # from tkinter import filedialog
        # filename =  filedialog.askopenfilename(initialdir="/",
        #             title = "Open file",
        #             filetypes = (("jpeg files", "*.jpg"),("all files", "*.*")))
        # filename = filedialog.asksaveasfilename(initialdir="/",
        #             title = "Save file",
        #             filetypes = (("jpeg files", "*.jpg"), ("all files", "*.*")))'''
        prt(line + "\n")

    # MESSAGE
    elif flds[wgt].lower() == "message":
        line = "{0} = Message({3}, text='{1}'{2})"
        prt(line.format(flds[var], flds[txt], attribs, flds[par]))
        line = "{0}.grid(row={1}, column={2}{3}{4}{5})\n"
        prt(line.format(flds[var], flds[row],
                        flds[col], rowspan, colspan, sticky))

    # POPUP MENU
    elif flds[wgt].lower().startswith("pop"):
        line = '''
        # self.popup_menu = Menu(self, tearoff=0)
        # self.popup_menu.add_command(label="Copy",
        #                             command=lambda:self.function(1))
        # self.popup_menu.add_command(label="Paste",
        #                             command=lambda:self.function(2))
        # self.popup_menu.add_separator()
        # self.popup_menu.add_command(label="say bye", command=exit)
        # self.txt.bind("<Button-3>", self.do_popup)

    # def do_popup(self,event):
    #     try:
    #         self.popup_menu.tk_popup(event.x_root,
    #                                  event.y_root)
    #     finally:
    #         self.popup_menu.grab_release()
        '''
        prt(line + "\n")

    # CLIPBOARD ACCESS
    elif flds[wgt].lower().startswith("clip"):
        line = '''
    # def clipbrd(self,n):
    #     if n == 1:  # Copy
    #         pyperclip.copy(self.txt.selection_get())
    #     else:
    #         # n == 2:  # Paste
    #         inx = self.txt.index(INSERT)
    #         self.txt.insert(inx, pyperclip.paste())
        '''
        prt(line + "\n")

    # GEOMETRY FOR WINDOW
    elif flds[wgt].lower().startswith("geo"):
        line = "root.geometry(\"%s\")" % (flds[var])
        prt(line + "\n")

    # SCALE
    elif flds[wgt].lower() == "scale":
        checkcols([2,3,5,6,7])
        prt("self." + flds[com] + " = DoubleVar()")
        line = "{0} = Scale({3}, variable=self.{1}{2}, from_=0, to=50)"
        prt(line.format(flds[var], flds[com], attribs, flds[par]))
        line = "{0}.grid(row={1}, column={2}{3}{4}{5})"
        prt(line.format(flds[var], flds[row],
                        flds[col], rowspan, colspan, sticky))
        line = '''# OTHER OPTIONS:
        # value=25,
        # orient=HORIZONTAL,
        # length=200,
        # command=self.scale_action
        #def scale_action(self, value):
            ## value = self.scale.get()  # or get current value
            # print(int(float(value)))
        '''
        prt(line + "\n")

    # TOPLEVEL
    elif flds[wgt].lower() == "toplevel":
        line = '''
    # def create_window(self):
    #     t = Toplevel(self)
    #     t.wm_title("Toplevel")
    #     t.geometry("200x100") # WxH+left+top
    #     l = Label(t, text="This is a Toplevel Window")
    #     l.grid(row=0, column=0, padx=2, pady=20)
    #     tvbtn = Button(t, text=" Exit ", command=t.destroy)
    #     tvbtn.grid(row=2, column=0, sticky='w', padx=2, pady=4)
    '''
        prt(line + "\n")


    # CALENDAR
    elif flds[wgt].lower() == "calendar":
        checkcols([2,3,6,7])
        line = "self.{0} = Calendar({2}{1})"
        prt(line.format(flds[var], attribs, flds[par]))
        line = "self.{0}.grid(row={1}, column={2}{3}{4}{5})\n"
        prt(line.format(flds[var], flds[row],
                        flds[col], rowspan, colspan, sticky))
        line = '''
        # self.cal = Calendar(self, selectmode="day",
        #                     year=int(strftime('%Y')),
        #                     month=int(strftime('%m')),
        #                     day=int(strftime('%d')),
        #                     width=400,
        #                     cursor="hand1",
        #                     date_pattern='yyyy-mm-dd')
        ###
        #  root.bind("<<CalendarSelected>>", self.calselected)
        #  self.calselected(None)
        #  date_key = self.cal.get_date()
        ###
        '''
        prt(line + "\n")


    else:
        if flds[wgt].startswith("Widget"):
            pass
        else:
            prt("\nNUNRECOGNIZED WIDGET IDENTIFIER: " + flds[wgt] + "\n\n")

    rownum += 1  # increment loop through spreadsheet


fout = open(outFile, "w")

# location of master templates
if args.bstrap:
    fin = open(TPLPATH + "tkbauto_tpl.py", "r")
else:
    fin = open(TPLPATH + "tkauto_tpl.py", "r")

for line in fin:
    if line.find("INSERT TKAUTO OUTPUT") > 0:
        fout.write(sout)
        #  insert the callbacks
        for item in callbacks:
            fout.write("    def %s(self):\n" % (item))
            fout.write("        ''' docstring '''\n\n")
    else:
        fout.write(line)


fout.close()
fin.close()
print(f"\n\nFind new script in {outFile}\n")

if errors is True:
    print("Some required options missing in one or more widgets!\n\n")
else:
    print("Some Widgets may need futher definition\n\n")

if args.exec:
    comline = "python3 " + outFile
    os.system(comline)
